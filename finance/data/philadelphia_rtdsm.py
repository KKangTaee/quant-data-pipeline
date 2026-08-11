"""Philadelphia Fed RTDSM full-history vintage ingestion."""

from __future__ import annotations

import json
import logging
import math
import re
import time
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, datetime, time as clock_time, timedelta, timezone
from io import BytesIO
from typing import Any, Callable, Iterable, Iterator, Sequence
from urllib.request import Request, urlopen
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from .db.mysql import MySQLClient
from .db.schema import PROVIDER_SCHEMAS, sync_table_schema
from .fred_vintages import (
    OPEN_ENDED_REALTIME_DATE,
    VINTAGE_TABLE,
    upsert_fred_vintage_rows,
)


DB_META = "finance_meta"
RTDSM_SOURCE = "philadelphia_fed_rtdsm"
DEFAULT_TIMEOUT = 90
DEFAULT_RETRIES = 3
DEFAULT_BATCH_SIZE = 5_000
LOGGER = logging.getLogger(__name__)


class RtdsmSourceError(RuntimeError):
    """Raised when an RTDSM workbook cannot satisfy the locked source contract."""


@dataclass(frozen=True)
class RtdsmSeriesSpec:
    """One official workbook and its research-only transform contract."""

    series_id: str
    workbook_url: str
    sheet_name: str
    series_name: str
    factor_group: str
    vintage_frequency: str
    units: str
    transform: str
    direction: int


_RTDMS_BASE = (
    "https://www.philadelphiafed.org/-/media/FRBP/Assets/Surveys-And-Data/"
    "real-time-data/data-files/xlsx"
)
_CATALOG = (
    RtdsmSeriesSpec(
        "IPT",
        f"{_RTDMS_BASE}/iptMvMd.xlsx",
        "ipt",
        "Total industrial production index",
        "activity",
        "monthly",
        "index_sa",
        "annualized_log_change_6m",
        1,
    ),
    RtdsmSeriesSpec(
        "H",
        f"{_RTDMS_BASE}/hMvMd.xlsx",
        "h",
        "Aggregate weekly hours index, total",
        "activity",
        "monthly",
        "index_sa",
        "annualized_log_change_3m",
        1,
    ),
    RtdsmSeriesSpec(
        "EMPLOY",
        f"{_RTDMS_BASE}/employMvMd.xlsx",
        "employ",
        "Nonfarm payroll employment",
        "labor_income",
        "monthly",
        "thousands_sa",
        "annualized_log_change_3m",
        1,
    ),
    RtdsmSeriesSpec(
        "RUC",
        f"{_RTDMS_BASE}/rucQvMd.xlsx",
        "ruc",
        "Unemployment rate",
        "labor_income",
        "quarterly",
        "percentage_points_sa",
        "level_change_3m",
        -1,
    ),
)


def get_rtdsm_catalog() -> tuple[RtdsmSeriesSpec, ...]:
    """Return the locked provider-native history catalog."""

    return _CATALOG


def _month_end(year: int, month: int) -> date:
    return date(year, month, monthrange(year, month)[1])


def _two_digit_year(value: str) -> int:
    year = int(value)
    return 2000 + year if year <= 40 else 1900 + year


def parse_rtdsm_vintage_header(
    header: object,
    *,
    series_id: str,
    vintage_frequency: str,
) -> date:
    """Map a provider header to a conservative month-end known-at date."""

    text = str(header or "").strip().upper()
    normalized_series = str(series_id or "").strip().upper()
    frequency = str(vintage_frequency or "").strip().lower()
    if frequency == "monthly":
        match = re.fullmatch(
            rf"{re.escape(normalized_series)}(\d{{2}})M(\d{{1,2}})",
            text,
        )
        if match is None:
            raise RtdsmSourceError(f"Invalid monthly vintage header: {header!r}")
        year = _two_digit_year(match.group(1))
        month = int(match.group(2))
    elif frequency == "quarterly":
        match = re.fullmatch(
            rf"{re.escape(normalized_series)}(\d{{2}})Q([1-4])",
            text,
        )
        if match is None:
            raise RtdsmSourceError(f"Invalid quarterly vintage header: {header!r}")
        year = _two_digit_year(match.group(1))
        month = {1: 2, 2: 5, 3: 8, 4: 11}[int(match.group(2))]
    else:
        raise RtdsmSourceError(f"Unsupported vintage frequency: {vintage_frequency!r}")
    try:
        return _month_end(year, month)
    except ValueError as exc:
        raise RtdsmSourceError(f"Invalid vintage header: {header!r}") from exc


def _observation_date(value: object) -> date:
    text = str(value or "").strip()
    match = re.fullmatch(r"(\d{4}):(\d{2})", text)
    if match is None:
        raise RtdsmSourceError(f"Invalid observation month: {value!r}")
    try:
        return date(int(match.group(1)), int(match.group(2)), 1)
    except ValueError as exc:
        raise RtdsmSourceError(f"Invalid observation month: {value!r}") from exc


def _collected_at_text(value: datetime | str) -> str:
    if isinstance(value, datetime):
        resolved = value
        if resolved.tzinfo is not None:
            resolved = resolved.astimezone(timezone.utc).replace(tzinfo=None)
        return resolved.strftime("%Y-%m-%d %H:%M:%S")
    text = str(value or "").strip()
    if not text:
        raise RtdsmSourceError("collected_at is required")
    return text


def _released_at(known_at: date) -> str:
    local = datetime.combine(
        known_at,
        clock_time(23, 59, 59, 999999),
        tzinfo=ZoneInfo("America/New_York"),
    )
    return local.astimezone(timezone.utc).isoformat()


def _finite_value(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if text in {"", ".", "-", "--", "#N/A", "NA", "N/A", "nan", "None"}:
        return None
    try:
        parsed = float(text.replace(",", ""))
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def _repair_xlsx_core_metadata(payload: bytes) -> BytesIO:
    """Repair only malformed single-digit XLSX metadata hours in memory."""

    output = BytesIO()
    try:
        with ZipFile(BytesIO(payload)) as source, ZipFile(
            output, "w", ZIP_DEFLATED
        ) as target:
            for member in source.infolist():
                content = source.read(member.filename)
                if member.filename == "docProps/core.xml":
                    content = re.sub(
                        rb"T ([0-9]):",
                        lambda match: b"T0" + match.group(1) + b":",
                        content,
                    )
                target.writestr(member, content)
    except BadZipFile as exc:
        raise RtdsmSourceError("RTDSM payload is not a valid XLSX archive") from exc
    output.seek(0)
    return output


def _selected_vintages(
    spec: RtdsmSeriesSpec,
    headers: Sequence[object],
    *,
    minimum_vintage_date: str | date | None,
) -> list[tuple[int, date, date]]:
    parsed = [
        parse_rtdsm_vintage_header(
            header,
            series_id=spec.series_id,
            vintage_frequency=spec.vintage_frequency,
        )
        for header in headers
    ]
    if len(parsed) != len(set(parsed)):
        raise RtdsmSourceError("RTDSM vintage headers must be unique")
    if parsed != sorted(parsed) or any(
        current <= previous for previous, current in zip(parsed, parsed[1:])
    ):
        raise RtdsmSourceError("RTDSM vintage headers must be strictly increasing")
    minimum = (
        date.fromisoformat(str(minimum_vintage_date)[:10])
        if minimum_vintage_date is not None
        else None
    )
    selected: list[tuple[int, date, date]] = []
    for index, known_at in enumerate(parsed):
        if minimum is not None and known_at < minimum:
            continue
        realtime_end = (
            parsed[index + 1] - timedelta(days=1)
            if index + 1 < len(parsed)
            else date.fromisoformat(OPEN_ENDED_REALTIME_DATE)
        )
        selected.append((index + 1, known_at, realtime_end))
    if not selected:
        raise RtdsmSourceError(
            f"No {spec.series_id} vintages satisfy the incremental lower bound"
        )
    return selected


def iter_rtdsm_normalized_batches(
    spec: RtdsmSeriesSpec,
    workbook_bytes: bytes,
    *,
    collected_at: datetime | str,
    minimum_vintage_date: str | date | None = None,
    batch_size: int = DEFAULT_BATCH_SIZE,
) -> Iterator[list[dict[str, object]]]:
    """Validate one wide workbook and stream normalized finite observations."""

    resolved_batch_size = int(batch_size)
    if resolved_batch_size < 1:
        raise RtdsmSourceError("batch_size must be at least 1")
    try:
        workbook = load_workbook(
            _repair_xlsx_core_metadata(workbook_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise RtdsmSourceError(f"Unable to open RTDSM workbook: {type(exc).__name__}") from exc
    if workbook.sheetnames != [spec.sheet_name]:
        raise RtdsmSourceError(
            f"RTDSM sheet contract mismatch: expected {spec.sheet_name!r}, "
            f"received {workbook.sheetnames!r}"
        )
    sheet = workbook[spec.sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise RtdsmSourceError("RTDSM workbook has no header row") from exc
    if not header_row or str(header_row[0] or "").strip().upper() != "DATE":
        raise RtdsmSourceError("RTDSM first header must be DATE")
    selected = _selected_vintages(
        spec,
        header_row[1:],
        minimum_vintage_date=minimum_vintage_date,
    )
    source_mode = f"rtdsm_full_history_{spec.vintage_frequency}"
    collected_text = _collected_at_text(collected_at)
    batch: list[dict[str, object]] = []
    emitted = 0
    for raw_row in rows:
        if not raw_row or raw_row[0] in (None, ""):
            continue
        observation = _observation_date(raw_row[0])
        for column_index, realtime_start, realtime_end in selected:
            value = (
                _finite_value(raw_row[column_index])
                if column_index < len(raw_row)
                else None
            )
            if value is None:
                continue
            emitted += 1
            batch.append(
                {
                    "series_id": spec.series_id,
                    "observation_date": observation.isoformat(),
                    "realtime_start": realtime_start.isoformat(),
                    "realtime_end": realtime_end.isoformat(),
                    "released_at": _released_at(realtime_start),
                    "source": RTDSM_SOURCE,
                    "source_type": "official",
                    "source_mode": source_mode,
                    "source_ref": spec.workbook_url,
                    "series_name": spec.series_name,
                    "factor_group": spec.factor_group,
                    "frequency": "monthly",
                    "units": spec.units,
                    "value": value,
                    "release_lag_days": (realtime_start - observation).days,
                    "coverage_status": "actual",
                    "missing_fields_json": json.dumps([]),
                    "collected_at": collected_text,
                    "error_msg": None,
                }
            )
            if len(batch) >= resolved_batch_size:
                yield batch
                batch = []
    if batch:
        yield batch
    if emitted == 0:
        raise RtdsmSourceError("RTDSM workbook contains no numeric observations")


def _request_workbook(url: str, timeout: int) -> bytes:
    request = Request(
        url,
        headers={
            "Accept": (
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            "User-Agent": "quant-data-pipeline/rtdsm-history-v1",
        },
    )
    with urlopen(request, timeout=int(timeout)) as response:
        return response.read()


def download_rtdsm_workbook(
    spec: RtdsmSeriesSpec,
    *,
    timeout: int = DEFAULT_TIMEOUT,
    retries: int = DEFAULT_RETRIES,
    request_fn: Callable[[str, int], bytes] = _request_workbook,
    sleep_fn: Callable[[float], None] = time.sleep,
) -> bytes:
    """Download one official workbook with bounded retries."""

    attempts = max(1, int(retries))
    last_error: Exception | None = None
    for attempt in range(attempts):
        try:
            payload = request_fn(spec.workbook_url, int(timeout))
            if not payload:
                raise RtdsmSourceError("RTDSM workbook response is empty")
            return payload
        except Exception as exc:
            last_error = exc
            if attempt + 1 < attempts:
                sleep_fn(0.4 * (attempt + 1))
    detail = type(last_error).__name__ if last_error is not None else "unknown"
    raise RtdsmSourceError(
        f"{spec.series_id} workbook download failed: {detail}"
    ) from None


def load_latest_rtdsm_realtime_starts(
    series_ids: Iterable[str],
    *,
    db: object,
) -> dict[str, str]:
    """Return latest stored RTDSM vintage per provider-native series."""

    requested = tuple(
        dict.fromkeys(str(item or "").strip().upper() for item in series_ids)
    )
    if not requested or any(not item for item in requested):
        raise RtdsmSourceError("At least one RTDSM series_id is required")
    placeholders = ",".join(["%s"] * len(requested))
    rows = db.query(
        f"""
        SELECT series_id, MAX(realtime_start) AS latest_realtime_start
        FROM {VINTAGE_TABLE}
        WHERE source = %s
          AND series_id IN ({placeholders})
        GROUP BY series_id
        """,
        (RTDSM_SOURCE, *requested),
    )
    return {
        str(row["series_id"]).strip().upper(): str(
            row["latest_realtime_start"]
        )[:10]
        for row in rows
        if row.get("series_id") and row.get("latest_realtime_start")
    }


def collect_rtdsm_history(
    *,
    series_ids: Iterable[str] | None = None,
    connection: Any = None,
    timeout: int = DEFAULT_TIMEOUT,
    retries: int = DEFAULT_RETRIES,
    batch_size: int = DEFAULT_BATCH_SIZE,
    workbook_fetcher: Callable[..., bytes] = download_rtdsm_workbook,
    latest_vintage_loader: Callable[..., dict[str, str]] = (
        load_latest_rtdsm_realtime_starts
    ),
    writer: Callable[..., int] = upsert_fred_vintage_rows,
) -> dict[str, object]:
    """Collect the locked RTDSM catalog into the shared vintage ledger."""

    catalog = {item.series_id: item for item in get_rtdsm_catalog()}
    requested = (
        list(catalog)
        if series_ids is None
        else list(
            dict.fromkeys(str(item or "").strip().upper() for item in series_ids)
        )
    )
    unsupported = [series_id for series_id in requested if series_id not in catalog]
    if unsupported:
        raise RtdsmSourceError(
            f"Unsupported RTDSM series: {', '.join(unsupported)}"
        )

    owns_connection = connection is None
    db = connection or MySQLClient("localhost", "root", "1234", 3306)
    collected_at = datetime.now(timezone.utc)
    failed: list[dict[str, str]] = []
    series_rows: dict[str, int] = {}
    coverage: dict[str, int] = {}
    stored = 0
    try:
        if owns_connection:
            db.use_db(DB_META)
            schema = PROVIDER_SCHEMAS[VINTAGE_TABLE]
            db.execute(schema)
            sync_table_schema(db, VINTAGE_TABLE, schema, DB_META)
        latest = latest_vintage_loader(requested, db=db)
        for series_id in requested:
            try:
                payload = workbook_fetcher(
                    catalog[series_id],
                    timeout=int(timeout),
                    retries=int(retries),
                )
                row_count = 0
                for batch in iter_rtdsm_normalized_batches(
                    catalog[series_id],
                    payload,
                    collected_at=collected_at,
                    minimum_vintage_date=latest.get(series_id),
                    batch_size=int(batch_size),
                ):
                    written = int(writer(batch, db=db))
                    stored += written
                    row_count += written
                    for row in batch:
                        status = str(row["coverage_status"])
                        coverage[status] = coverage.get(status, 0) + 1
                series_rows[series_id] = row_count
            except Exception as exc:
                reason = str(exc)[:500]
                LOGGER.warning("RTDSM collection failed for %s: %s", series_id, reason)
                failed.append({"series_id": series_id, "reason": reason})
    finally:
        if owns_connection:
            db.close()

    missing = sorted(set(requested) - set(series_rows))
    return {
        "requested": len(requested),
        "stored": stored,
        "missing": missing,
        "failed": failed,
        "coverage": coverage,
        "series_rows": series_rows,
        "source": RTDSM_SOURCE,
        "source_mode": "rtdsm_full_history",
    }
