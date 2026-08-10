"""Collect official Philadelphia Fed SPF Core PCE Q4/Q4 probabilities."""

from __future__ import annotations

from datetime import date, datetime, time, timezone
from io import BytesIO
import re
from typing import Callable, Mapping
from urllib.request import Request, urlopen
from zipfile import ZIP_DEFLATED, ZipFile
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from .db.mysql import MySQLClient
from .db.schema import INFLATION_POLICY_SCHEMAS, sync_table_schema


DB_META = "finance_meta"
TABLE = "spf_core_pce_probability"
SPF_RELEASE_DATES_URL = (
    "https://www.philadelphiafed.org/-/media/frbp/assets/surveys-and-data/"
    "survey-of-professional-forecasters/spf-release-dates.txt"
)
SPF_CORE_PCE_MEAN_URL = (
    "https://www.philadelphiafed.org/-/media/FRBP/Assets/Surveys-And-Data/"
    "survey-of-professional-forecasters/data-files/files/Mean_PRCPCE_Level.xlsx"
)
PARSER_VERSION = "philadelphia-fed-spf-prcpce-v1"
RELEASE_TIME_BASIS = "official_release_date_end_of_day_america_new_york"

_BINS = (
    (1, ">=4.0", 4.0, None),
    (2, "3.5-3.9", 3.5, 3.9),
    (3, "3.0-3.4", 3.0, 3.4),
    (4, "2.5-2.9", 2.5, 2.9),
    (5, "2.0-2.4", 2.0, 2.4),
    (6, "1.5-1.9", 1.5, 1.9),
    (7, "1.0-1.4", 1.0, 1.4),
    (8, "0.5-0.9", 0.5, 0.9),
    (9, "0.0-0.4", 0.0, 0.4),
    (10, "decline", None, 0.0),
)


def _sql_datetime(value: datetime) -> str:
    return value.astimezone(timezone.utc).replace(tzinfo=None).strftime(
        "%Y-%m-%d %H:%M:%S.%f"
    )


def _aware_utc(value: str | datetime) -> datetime:
    parsed = (
        value
        if isinstance(value, datetime)
        else datetime.fromisoformat(str(value).strip().replace("Z", "+00:00"))
    )
    if parsed.tzinfo is None:
        raise ValueError("collected_at must include a timezone")
    return parsed.astimezone(timezone.utc)


def parse_spf_release_dates(text: str) -> dict[tuple[int, int], str]:
    """Parse the survey's official news-release date, carrying repeated years."""

    releases: dict[tuple[int, int], str] = {}
    current_year: int | None = None
    pattern = re.compile(
        r"^\s*(?:(20\d{2}|19\d{2})\s+)?Q([1-4])\s+"
        r"(\d{1,2}/\d{1,2}/\d{2})\*{0,3}\s+"
        r"(\d{1,2}/\d{1,2}/\d{2})\*{0,3}\s*$"
    )
    for line in str(text or "").splitlines():
        match = pattern.match(line)
        if match is None:
            continue
        if match.group(1):
            current_year = int(match.group(1))
        if current_year is None:
            continue
        quarter = int(match.group(2))
        released = datetime.strptime(match.group(4), "%m/%d/%y").date()
        releases[(current_year, quarter)] = released.isoformat()
    if not releases:
        raise ValueError("SPF release-date file contains no survey rows")
    return releases


def _repair_openxml_core_properties(payload: bytes) -> bytes:
    """Repair the official workbook's non-zero-padded W3CDTF hour in memory."""

    with ZipFile(BytesIO(payload)) as source:
        repaired = BytesIO()
        with ZipFile(repaired, "w", ZIP_DEFLATED) as target:
            for info in source.infolist():
                content = source.read(info.filename)
                if info.filename == "docProps/core.xml":
                    content = re.sub(
                        rb"T\s+([0-9]{1,2}):",
                        lambda match: b"T" + match.group(1).zfill(2) + b":",
                        content,
                    )
                target.writestr(info, content)
    return repaired.getvalue()


def _conservative_release_at(release_date: date) -> datetime:
    eastern = ZoneInfo("America/New_York")
    return datetime.combine(release_date, time.max, tzinfo=eastern).astimezone(
        timezone.utc
    )


def _missing_probability(value: object) -> bool:
    return value in (None, "") or str(value).strip().upper() in {
        "#N/A",
        "N/A",
        "NA",
    }


def parse_spf_core_pce_workbook(
    payload: bytes,
    *,
    release_dates: Mapping[tuple[int, int], str],
    collected_at: str | datetime,
    source_ref: str,
) -> list[dict[str, object]]:
    """Normalize current- and next-year PRCPCE mean probability distributions."""

    workbook = load_workbook(
        BytesIO(_repair_openxml_core_properties(bytes(payload))),
        read_only=True,
        data_only=True,
    )
    if "Mean_Level" not in workbook.sheetnames:
        raise ValueError("SPF workbook is missing Mean_Level")
    sheet = workbook["Mean_Level"]
    values = sheet.iter_rows(values_only=True)
    try:
        header = tuple(str(item or "").strip().upper() for item in next(values))
    except StopIteration as exc:
        raise ValueError("SPF workbook is empty") from exc
    expected = ("YEAR", "QUARTER", *(f"PRCPCE{i}" for i in range(1, 21)))
    if header != expected:
        raise ValueError("SPF workbook columns do not match the PRCPCE contract")
    collected = _aware_utc(collected_at)
    rows: list[dict[str, object]] = []
    for raw in values:
        if raw[0] in (None, "") or raw[1] in (None, ""):
            continue
        survey_year = int(float(raw[0]))
        survey_quarter = int(float(raw[1]))
        if not 1 <= survey_quarter <= 4:
            raise ValueError("SPF survey quarter must be between one and four")
        release_text = release_dates.get((survey_year, survey_quarter))
        if release_text is None:
            continue
        release_date = date.fromisoformat(str(release_text)[:10])
        released_at = _conservative_release_at(release_date)
        probabilities = raw[2:22]
        for horizon_index, horizon in enumerate(("CURRENT_YEAR", "NEXT_YEAR")):
            horizon_values = probabilities[horizon_index * 10 : (horizon_index + 1) * 10]
            if all(_missing_probability(value) for value in horizon_values):
                continue
            if any(_missing_probability(value) for value in horizon_values):
                raise ValueError("SPF probability horizon is only partially populated")
            parsed = [float(value) for value in horizon_values]
            if any(value < 0.0 or value > 100.0 for value in parsed):
                raise ValueError("SPF probabilities must be between zero and 100")
            if abs(sum(parsed) - 100.0) > 0.25:
                raise ValueError("SPF probability bins must sum to 100")
            target_year = survey_year + horizon_index
            for (bin_number, label, lower, upper), probability in zip(
                _BINS, parsed, strict=True
            ):
                rows.append(
                    {
                        "survey_year": survey_year,
                        "survey_quarter": survey_quarter,
                        "target_year": target_year,
                        "horizon": horizon,
                        "bin_number": bin_number,
                        "bin_label": label,
                        "bin_lower_pct": lower,
                        "bin_upper_pct": upper,
                        "mean_probability_pct": probability,
                        "release_date": release_date.isoformat(),
                        "released_at": _sql_datetime(released_at),
                        "release_time_basis": RELEASE_TIME_BASIS,
                        "source": "philadelphia_fed_spf",
                        "source_ref": str(source_ref),
                        "parser_version": PARSER_VERSION,
                        "collected_at": _sql_datetime(collected),
                    }
                )
    if not rows:
        raise ValueError("SPF workbook contains no released Core PCE distributions")
    return rows


def _fetch(url: str) -> bytes:
    request = Request(
        url,
        headers={"User-Agent": "quant-data-pipeline/1.0 (+macro research)"},
    )
    with urlopen(request, timeout=30) as response:
        return response.read()


def collect_and_store_spf_core_pce_probabilities(
    *,
    connection: object | None = None,
    fetch_text: Callable[[str], str] | None = None,
    fetch_bytes: Callable[[str], bytes] | None = None,
    collected_at: str | datetime | None = None,
    host: str = "localhost",
    user: str = "root",
    password: str = "1234",
    port: int = 3306,
) -> dict[str, object]:
    """Fetch and UPSERT all official historical SPF Core PCE probability bins."""

    text_loader = fetch_text or (lambda url: _fetch(url).decode("utf-8"))
    bytes_loader = fetch_bytes or _fetch
    releases = parse_spf_release_dates(text_loader(SPF_RELEASE_DATES_URL))
    rows = parse_spf_core_pce_workbook(
        bytes_loader(SPF_CORE_PCE_MEAN_URL),
        release_dates=releases,
        collected_at=collected_at or datetime.now(timezone.utc),
        source_ref=SPF_CORE_PCE_MEAN_URL,
    )
    owns_connection = connection is None
    db = connection or MySQLClient(host, user, password, port)
    try:
        db.use_db(DB_META)
        schema = INFLATION_POLICY_SCHEMAS[TABLE]
        db.execute(schema)
        sync_table_schema(db, TABLE, schema, DB_META)
        db.executemany(
            """
            INSERT INTO spf_core_pce_probability (
              survey_year, survey_quarter, target_year, horizon,
              bin_number, bin_label, bin_lower_pct, bin_upper_pct,
              mean_probability_pct, release_date, released_at,
              release_time_basis, source, source_ref, parser_version, collected_at
            ) VALUES (
              %(survey_year)s, %(survey_quarter)s, %(target_year)s, %(horizon)s,
              %(bin_number)s, %(bin_label)s, %(bin_lower_pct)s, %(bin_upper_pct)s,
              %(mean_probability_pct)s, %(release_date)s, %(released_at)s,
              %(release_time_basis)s, %(source)s, %(source_ref)s,
              %(parser_version)s, %(collected_at)s
            )
            ON DUPLICATE KEY UPDATE
              horizon = VALUES(horizon),
              bin_label = VALUES(bin_label),
              bin_lower_pct = VALUES(bin_lower_pct),
              bin_upper_pct = VALUES(bin_upper_pct),
              mean_probability_pct = VALUES(mean_probability_pct),
              release_date = VALUES(release_date),
              released_at = VALUES(released_at),
              release_time_basis = VALUES(release_time_basis),
              source = VALUES(source),
              source_ref = VALUES(source_ref),
              parser_version = VALUES(parser_version),
              collected_at = VALUES(collected_at)
            """,
            rows,
        )
    finally:
        if owns_connection:
            db.close()
    latest = max((int(row["survey_year"]), int(row["survey_quarter"])) for row in rows)
    return {
        "status": "success",
        "stored": len(rows),
        "rows": len(rows),
        "latest_survey": f"{latest[0]}Q{latest[1]}",
        "source": "philadelphia_fed_spf",
    }
